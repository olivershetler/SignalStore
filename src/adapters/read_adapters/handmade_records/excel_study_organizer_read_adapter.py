from src.operations.importers.adapters.abstract_read_adapter import AbstractReadAdapter

import openpyxl as xl

class ExcelStudyOrganizerReadAdapter(AbstractReadAdapter):
    def __init__(self, path):
        self.path = path
        self.wb = xl.load_workbook(path)
        self.ws = self.wb.active
        self.tables = [str(table) for table in self.wb.sheetnames]

    def read(self):
        for table in self.tables:
            for record in self._get_table_records(table):
                yield record

    def read_records(self):
        records = []
        for table in self.tables:
            records.extend(list(self._get_table_records(table)))
        return records

    def read_records_by_table(self):
        records = {}
        for table in self.tables:
            records[str(table).lower()] = list(self._get_table_records(table))
        return records

    def _classify_table(self, table):
        # check if name column is unique
        has_unique_keys = self._has_unique_keys(table)

        # check if the table columns include only the name column,
        # an attribute column and a value column
        is_attr_value_format = self._is_attr_value_format(table)

        if not has_unique_keys and is_attr_value_format:
            return 'attribute'
        elif has_unique_keys and not is_attr_value_format:
            return 'record'
        else:
            error_string = f'Could not classify table {table}.'
            if not has_unique_keys:
                error_string += '\nTable does not have unique keys.'
            if not is_attr_value_format:
                error_string += '\nTable is not in attribute-value format.'
            raise StudyOrganizerKeyError(error_string)

    def _has_unique_keys(self, table):
        ws = self.wb[table]
        keys = [str(cell.value).lower() for cell in ws['A'] if cell.value is not None]
        return len(keys) == len(set(keys))

    def _is_attr_value_format(self, table):
        ws = self.wb[table]
        columns = [str(cell.value).lower() for cell in ws[1]]
        if columns[0] == f'name' and columns[1] == 'attribute' and columns[2] == 'value' and len(columns) == 3:
            return True
        else:
            return False

    def _get_table_records(self, table):
        table_type = self._classify_table(table)
        readers = {'record': self._get_simple_table_records,
                   'attribute': self._get_attribute_table_records}
        records = readers[table_type](table)
        for record in records:
            yield record

    def _get_simple_table_records(self, table):
        self.ws = self.wb[table]
        columns = [str(cell.value).lower() for cell in self.ws[1]]
        self._validate_columns(columns, table)
        for row in self.ws.iter_rows(min_row=2):
            record = {}
            for i, column in enumerate(columns):
                value = row[i].value
                if value is not None:
                    record[column] = value
            if record != {}:
                record['type'] = table
                yield record

    def _get_attribute_table_records(self, table):
        self.ws = self.wb[table]
        columns = [str(cell.value).lower() for cell in self.ws[1]]
        self._validate_columns(columns, table)
        attr_records = []
        for row in self.ws.iter_rows(min_row=2):
            record = {}
            for i, column in enumerate(columns):
                record[column] = row[i].value
            attr_records.append(record)
        records = {}
        for attr_record in attr_records:
            rkey = attr_record['name']
            if rkey not in records.keys() and rkey is not None:
                records[rkey] = {'name': rkey, 'type': table}
            if attr_record['value'] is not None:
                records[rkey][attr_record['attribute']] = attr_record['value']
        for record in records.values():
            if record != {}:
                yield record

    def _validate_columns(self, columns, table_name):
        if not self._first_column_is_key(columns, table_name):
            raise StudyOrganizerKeyError(f'First column must be a "name" column, but is {columns[0]}.')

    def _first_column_is_key(self, columns, table_name):
        return str(columns[0]) == f'name'


class StudyOrganizerKeyError(KeyError):
    pass